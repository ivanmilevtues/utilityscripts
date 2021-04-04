from os import path
from openpyxl import Workbook
from openpyxl.chart import (ScatterChart, Reference, Series)



folder = "results/"

file_prefix = "simulation"

file_chunks = [64, 64*2, 64*3]

item_columns = '"Item";"Cycles";"Delivered packets";"Received packets";"Delivery rate (flits/node/cycle)";"Arrival rate (flits/node/cycle)";"Head Latency";"Head Latency (no locks)";"Packet Latency";"Packet Latency (no locks)";"Lock cycles";"Average Path (nodes)"\n'

def generate_chunks():
    for chunk in file_chunks:
        indx = file_chunks.index(chunk)
        f = open(folder + "chunk_file_" + str(indx) + ".csv", 'w')
        f.write(item_columns)
        start = 0 if indx == 0 else file_chunks[indx - 1]
        end = file_chunks[indx]

        concatenate_to_chunk(f, start, end)
        f.close()


def concatenate_to_chunk(file, start_chunk, end_chunk):
    for index in range(start_chunk + 1, end_chunk + 1):
        file_name = folder + file_prefix + ("0" if index < 10 else "") + str(index) + ".csv"
        if not path.exists(file_name):
            continue
        with open(file_name, 'r') as f:
            contents = f.readlines()
            for content in contents[1:]:
                file.write(content)

def get_file_name(index):
    return folder + file_prefix + ("0" if index < 10 else "") + str(index) + ".csv"



def generated_ex_one_csv(algo_id, buf_id, chan_id, pack_size):
    simulation_id = 0
    # print("Algorithm={0}, buffer={1}, channel={2}, package={3}".format(algorithms[algo_id], buffers[buf_id], in_channel[chan_id], package_sizes[pack_size]))
    columns = []
    for algorithm_id in range (3):
        for buffer_id in range(4):
            for channel_id in range(4):
                for package_size in range(4):
                    simulation_id += 1
                    if buffer_id == buf_id and algorithm_id==algo_id and pack_size == package_size:
                        # print("Algorithm={0}, buffer={1}, channel={2}, package={3}".format(algorithms[algorithm_id], buffers[buf_id], in_channel[chan_id], package_sizes[pack_size]))
                        filename = get_file_name(simulation_id)
                        avrg = 0
                        if not path.exists(filename):
                            continue
                        with open(filename, 'r') as f:
                            print(simulation_id)
                            contents = f.readlines()
                            latency_index = 7
                            try:
                                latency_index = contents[0].split(';').index('"Packet Latency"')
                            except ValueError:
                                print("error")
                            for content in contents[1:]:
                                # print('|{0}|'.format(content.split(';')[latency_index].replace(',', '.')))
                                # print(content.split(';')[latency_index].replace(',', '.'))
                                # print(float(content.split(';')[latency_index].replace(',', '.')))
                                avrg += float(content.split(';')[latency_index].replace(',', '.'))

                            # print(avrg, len(contents))
                            lenght = len(contents) - 1
                            if lenght != 0:
                                columns.append([int(in_channel[buf_id]), avrg / lenght])
    return columns

algorithms = ['Dimension order (XY) for meshes (deterministic)', 'Duato based on (XY) for meshes (adaptive)', 'Fully adaptive for meshes(with deadlocks)']

buffers = ['1', '2', '4', '8']

in_channel = ['1', '2', '4', '8']

package_sizes = ['32', '64', '128', '256']


def do_exercize():
    workbook = Workbook()
    for i in range(0, 4):
        for k in range(0, 4):
            columns = generated_ex_one_csv(algo_id=1, buf_id=i, chan_id=0, pack_size=k)
            print(columns)
            append_to_workbook("BufferSize={0},PackageSize={1}".format(buffers[i], package_sizes[k]), workbook, columns)
    workbook.save("exersize_5_2.xlsx")

def do_exercize2():
    workbook = Workbook()

    columns = generated_ex_one_csv(algo_id=1, buf_id=None, chan_id=0, pack_size=None)
    print(columns)
    append_to_workbook("Injection{0}".format(in_channel[0]), workbook, columns)
    columns = generated_ex_one_csv(algo_id=1, buf_id=None, chan_id=2, pack_size=None)
    append_to_workbook("Injection{0}".format(in_channel[2]), workbook, columns)
    workbook.save("exersize_4.xlsx")

def append_to_workbook(sheet_name, workbook: Workbook, column_info):
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["Buffer size", "Packet latency (adaptive)"])
    for info in column_info:
        sheet.append(info)
    
    # chart = ScatterChart()
    # chart.legend.position = 'b'
    # chart.title = 'Packet latency = f(Injection buffer)'
    # chart.x_axis.title = 'Injection buffer'
    # chart.y_axis.title = 'Packet latency'

    # x_values = Reference(worksheet=sheet,
    #                 min_row=2,
    #                 max_row=len(column_info) + 1,
    #                 min_col=1)
    # y_values = Reference(worksheet=sheet,
    #             min_row=2,
    #             max_row=len(column_info) + 1,
    #             min_col=2)
    # series = Series(y_values, x_values, title=sheet_name)
    # chart.series.append(series)
    # sheet.add_chart(chart, 'E2')

do_exercize()